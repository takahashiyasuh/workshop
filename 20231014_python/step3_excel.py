# ライブラリのインポート
import mysql.connector
import openpyxl

# DB情報
DB_HOST='127.0.0.1'
DB_USER='root'
DB_PASSWORD='password'
DB_NAME='sales'

conn = None
curs = None

# DBコネクション
conn = mysql.connector.connect(host=DB_HOST,
                            user=DB_USER,
                            password=DB_PASSWORD,
                            database=DB_NAME)

# カーソル取得
curs = conn.cursor()

# データを取得
sql = 'SELECT id, goods, price, buy_date FROM sales;'
curs.execute(sql)
# fetchall()は中身を全て取得するメソッド
rows = curs.fetchall()

# 取得したデータを表示
for id, goods, price, buy_date in rows:
    print(id, goods, price, buy_date)

# 終了処理
if curs is not None:
    curs.close()
if conn is not None:
    conn.close()

# ワークブックを作成
wb = openpyxl.Workbook()

# 既存のシートを削除
wb.remove(wb['Sheet'])

# ワークシートを作成
ws = wb.create_sheet(title='sales')

# 取得したデータを表示
for index, (id, goods, price, buy_date) in enumerate(rows):
    ws.cell(row=index+2, column=1).value = id
    ws.cell(row=index+2, column=2).value = goods
    ws.cell(row=index+2, column=3).value = price
    ws.cell(row=index+2, column=4).value = buy_date

# 円グラフのための表を作成
dict = {}
for (id, goods, price, buy_date) in rows:
    if goods in dict:
        # 個数を増加
        dict[goods] += 1
    else:
        # ディクショナリに追加
        dict[goods] = 1

# 集計ワークシートを新規作成
ws = wb.create_sheet(title='goods')

# 表を作成する
for index, (key, value) in enumerate(dict.items()):
    ws.cell(row=index+2, column=1).value = key
    ws.cell(row=index+2, column=2).value = value

# ワークブックを保存
wb.save('sales.xlsx')

# 閉じる
wb.close()